from datetime import datetime
from openpyxl import load_workbook
rut=r"C:\Users\SENA\Desktop\base_crud.xlsx"





def leer(ruta:str. extraer:str):
    arcrivo_exccel = load_workbook("C:\Users\SENA\Desktop\base_crud.xlsx")
    hoja_datos = archivo_Exccel['Datos del crud']
    hoja_datos=hoja_datos['A2':'F'+str(hoja_datos.max_row)]

    info={}

    for i in hoja_datos:

        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{'tarea':i[1].value, 'descripcion':i[2].value,
                                        'estado':i[3].value, 'fecha de inicio':i[4].value,
                                        'fecha finalizacon':i[5].value})
    if not(estraer=='todo'):
        info=filtrar(info, extraer)
    
    for i in info:
        print('******** Tarea *******')
        print('Id'+ str(i)+'\n'+'Titulo: '+str(info[i]['tarea'])+'\n'+'Descripcion: '
              +str(info[i]['descripcion']) + '\n'+ 'Estado:'+str(info[i]['Estado'])
              +'\n'+'Fecha Creacion: '+ str(info[i]['fecha de inicio'])
              + '\n'+ 'Fecha de finalizacion: ' +str(info[i]['fecha de finalizacion']) )
        print()

    return

def filtrar(info:dict,filtro:str):
    aux={}

    for i in info:
        if info[i]['estado']==filtro:
            aux.setdefault(i,info[i])
    
    return aux

def actualizar(ruta:str, identicador:int,datos_acualizados:dict):
    Archivo_Exccel = load_worbook("C:\Users\SENA\Desktop\base_crud.xlsx")
    Hoja_datos = Archivo_Exccel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    Hoja=Archivo_Exccel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizacion=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            for d in datos_actualizaados:
                if d=='titulo' and not(datosActualzados[d]==''):
                hoja.cell(row=fila, column=titulo).value=datosActualizados[d]
                elif d=='descripcion' and not(datosActualizados[d]==''):
                    hoja.cell(row=fila, column=descripcion).value=datosActualizados[d]
                elif d=='estado' and not(datosActualizados[d]==''):
                    hoja.cell(row=fila, column=estado).value=datosActualizados[d]
                elif d=='fecha inicio' and not(datosActualizados[d]==''):
                    hoja.cell(row=fila, column=fecha inicio)value=datosActualizados[d]
                elif d=='fecha finalizacion' and not(datosActualizados[d]==''):
                    hoja.cell(row=fila, column=fecha finalizacion)value=datosActualizados[d]
    Archivo_Exccel.save("C:\Users\SENA\Desktop\base_crud.xlsx")
    if encontro==False:
        print('Error. No existe una tarea con ese Id')
        print()
    return

def agregar(ruta:int, datos:dict):
    Archivo_Exccel = load_workbook("C:\Users\SENA\Desktop\base_crud.xlsx")
    Hoja_datos = Archivo_Exccel['Datos del crud']
    Hoja_datos=hoja_datos['A2':'F'+str(Hoja_datos.max_row+1)]
    Hoja=Archivo_Exccel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizacion=6
    for i in Hoja_datos:

        if not(isinstance(i[0].value, int)):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=titulo).value=datos['titulo']
            hoja.cell(row=identificador, column=descripcion).value=datos['descripcion']
            hoja.cell(row=identificador, column=estado).value=['estado']
            hoja.cell(row=identificador, column=fecha_inicio).value=datos['fecha inicio']
            hoja.cell(row=identificador, column=fecha_finalizacion).value=datos['fecha finalizacion']
            break
    Archivo_Exccel.save("C:\Users\SENA\Desktop\base_crud.xlsx")
    return

def borrar(ruta,identificador):
    Archivo_Exccel = load_workbook("C:\Users\SENA\Desktop\base_crud.xlsx")
    Hoja_datos = Archivo_Exccel['Datos del crud']
    Hoja_datos=hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    Hoja=Archivo_Exccel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizacion=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            hoja.cell(row=fila, column=1).value=""
            hoja.cell(row=fila, column=titulo).value=""
            hoja.cell(row=fila, column=descripcion).value=""
            hoja.cell(row=fila, column=estado).value=""
            hoja.cell(row=fila, column=fecha_inicio).value=""
            hoja.cell(row=fila, column=fecha_finalizacion).value=""
    Archivo_Exccel.save("C:\Users\SENA\Desktop\base_crud.xlsx")
    if encontro==False:
        print('Error: No existe ua tarea con ese Id')
        print()
    return


rut="C:\Users\SENA\Desktop\base_crud.xlsx"

datosActualizados={'titulo':'', 'descripcion':'', 'estado':'','fecha inicio':'','fecha finalizacion'}
while True:
    print('Indique la accion que desea realizar: ')
    print('Consultar: 1')
    print('Actualizar: 2')
    print('Crear nueva tarea: 3')
    print('Borrar: 4')
    accion = input('Escriba la opcion: ')

if not(accion=='1') and not(acccion=='2') and not(accion=='3') and not(accion=='4'):
    print('Comando invalido por favir eliga una opcion valida')
elif accion=='1':
    opc_consulta=''
    print('Indique la tarea que desea consultar: ')
    print('Todas las tareas: 1')
    print('En espera: 2')
    print('En ejecucion: 3')
    print('Por aprobar: 4')
    print('Finalizada: 5')
    opc_consulta = input('Esvriba la tarea que desea consultar: ')
    if opc_consulta=='1':
        print()
        print()
        print('** Colsultando tareas en espera **')
        leer(rut,'todo')
    elif opc_consulta=='2':
        print()
        print()
        print('** Consultando tareas en espera **')
        leer(rut,'En espera')
    elif opc_consultando=='3':
        print()
        print()
        print('** Consultando tareas en ejecucion **')
        leer(rut, 'En ejecucuion')
    elif opc_consulta=='4':
        print()
        print()
        print('** Consultando tarea por aprobar **')
        leer(rtu,'Por aprobar')
    elif opc_consulta=='5':
        print()
        print()
        print('** Consultando tareas finalizadas **')
        leer(rut'finalizada')

